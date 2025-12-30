import requests
import pandas as pd
from datetime import datetime, timedelta
import time
import shutil
import os
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook

# 接口2配置（预约数据）
URL_DETAIL = 'https://channels.weixin.qq.com/micro/statistic/cgi-bin/mmfinderassistant-bin/statistic/live_single_data'

# 接口3配置（带货商品的数据）
URL_PRODUCT = 'https://channels.weixin.qq.com/micro/statistic/cgi-bin/mmfinderassistant-bin/statistic/get_single_live_ec_spu_data_page_v2'

# 接口1配置（列表数据）
# 已将请求 URL 固定为下面的值
URL_LIST = 'https://channels.weixin.qq.com/micro/statistic/cgi-bin/mmfinderassistant-bin/live/get_live_history'

# 接口4配置（带货数据的整体转换数据）
URL_EC_SUMMARY = 'https://channels.weixin.qq.com/micro/statistic/cgi-bin/mmfinderassistant-bin/statistic/get_single_live_funnel'

# 接口5配置（数据增强）
URL_DIAGNOSTIC = 'https://channels.weixin.qq.com/micro/statistic/cgi-bin/mmfinderassistant-bin/svrkit/MMFinderLiveDashboardDatasvr/getLiveDiagnosticData'

# HEADERS 直接写死为空（可由浏览器会话覆盖）
HEADERS = {}

# COOKIES 直接写死为空（可由浏览器会话覆盖）
COOKIES_DICT = {}


 
# 浏览器 profile 目录（用于从持久化上下文读取 cookies / UA）
BROWSER_USER_DATA_DIR = './browser_data'

def get_browser_session_cookies_and_headers(user_data_dir=BROWSER_USER_DATA_DIR, url=None):
    """
    从 Playwright 的持久化上下文读取 cookies 和 User-Agent，返回 (headers_dict, cookies_dict)
    如果失败返回 ({}, {})
    """
    try:
        playwright = sync_playwright().start()
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            headless=True,
        )

        pages = context.pages
        page = pages[0] if pages else context.new_page()

        # 如果提供了 url，尝试打开以便获得对应域的 cookies
        if url:
            try:
                page.goto(url, wait_until='domcontentloaded', timeout=5000)
            except:
                pass

        # 获取 cookies 列表并转换为 dict
        try:
            cookies_list = context.cookies()
            cookies = {c.get('name'): c.get('value') for c in cookies_list}
        except:
            cookies = {}

        # 获取 User-Agent
        try:
            ua = page.evaluate("() => navigator.userAgent")
            headers = {}
            if ua:
                headers['User-Agent'] = ua
        except:
            headers = {}

        try:
            context.close()
        except:
            pass
        try:
            playwright.stop()
        except:
            pass

        return headers, cookies
    except Exception as e:
        print(f"  [Warning] 从浏览器会话获取 cookies/headers 失败: {e}")
        try:
            playwright.stop()
        except:
            pass
        return {}, {}

def get_time_range_for_half_year(start_date_str=None, end_date_str=None):
    """获取时间范围

    Args:
        start_date_str: 开始日期字符串，格式为 'YYYY-MM-DD'，默认为今年1月1号
        end_date_str: 结束日期字符串，格式为 'YYYY-MM-DD'，默认为当前时间

    Returns:
        tuple: (start_time, end_time) 时间戳
    """
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            start_time = int(start_date.timestamp())
        except ValueError:
            print(f"无效的开始日期格式: {start_date_str}，使用默认值")
            start_date = datetime(2025, 1, 1, 0, 0, 0)
            start_time = int(start_date.timestamp())
    else:
        # 默认从今年1月1号开始
        start_date = datetime(2025, 1, 1, 0, 0, 0)
        start_time = int(start_date.timestamp())

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str + ' 23:59:59', '%Y-%m-%d %H:%M:%S')
            end_time = int(end_date.timestamp())
        except ValueError:
            print(f"无效的结束日期格式: {end_date_str}，使用当前时间")
            end_time = int(time.time())
    else:
        end_time = int(time.time())  # 当前时间戳

    return start_time, end_time

def fetch_live_data(page_size=10, current_page=1, start_time=None, end_time=None, headers=None, cookies=None):
    """获取直播列表数据（接口1）

    Args:
        page_size: 每页数据条数，默认10
        current_page: 当前页码，默认1
        start_time: 开始时间戳，如果为None则使用默认时间范围
        end_time: 结束时间戳，如果为None则使用默认时间范围
        headers: 请求头，如果为None则使用默认headers
        cookies: cookies字典，如果为None则使用默认cookies

    Returns:
        dict: API响应数据，包含直播列表信息或None（请求失败时）
    """
    if start_time is None or end_time is None:
        start_time, end_time = get_time_range_for_half_year()
    
    payload = {
        "pageSize": page_size,
        "currentPage": current_page,
        "reqType": 2,
        "filterStartTime": start_time,
        "filterEndTime": end_time,
        "timestamp": str(int(time.time() * 1000)),
        "rawKeyBuff": None,
        "pluginSessionId": None,
        "scene": 7,
        "reqScene": 7
    }
    
    # 优先使用传入的 headers/cookies（来自浏览器会话），否则回退到写死的 HEADERS/COOKIES_DICT
    request_headers = headers if headers is not None and headers else HEADERS
    request_cookies = cookies if cookies is not None and cookies else COOKIES_DICT

    try:
        response = requests.post(
            URL_LIST,
            json=payload,
            headers=request_headers,
            cookies=request_cookies,
            timeout=10
        )
        response.raise_for_status()
        data = response.json()
        
        if data.get('errCode') == 0:
            return data.get('data', {})
        else:
            print(f"API error: {data.get('errMsg')}")
            return None
    except Exception as e:
        print(f"Request error: {e}")
        return None

def save_records_to_excel_file(output_file, all_records, sheet_name='产品数据', id_column_name='liveobjectid', silent=False):
    """保存所有记录到Excel文件（覆盖写入，不追加）

    Args:
        output_file: Excel文件路径
        all_records: 所有记录字典列表（包含之前已保存的记录和新记录）
        sheet_name: 工作表名称
        id_column_name: ID列名称（用于设置文本格式）
        silent: 是否静默模式（不输出日志）

    Returns:
        bool: 是否成功保存
    """
    try:
        if not all_records:
            return True

        # 确保所有记录的 ID 列是字符串格式
        for record in all_records:
            if id_column_name in record:
                record[id_column_name] = str(record[id_column_name])

        # 创建DataFrame
        df = pd.DataFrame(all_records)

        # 确保 ID 列是字符串类型
        if id_column_name in df.columns:
            df[id_column_name] = df[id_column_name].astype(str)

        # 保存到Excel文件，使用 openpyxl 引擎以便后续设置格式
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            # 获取工作表并设置 ID 列为文本格式
            ws = writer.sheets[sheet_name]
            # 查找 ID 列的索引
            header_row = 1
            id_col = None
            for col_idx, cell in enumerate(ws[header_row], 1):
                if cell.value == id_column_name:
                    id_col = col_idx
                    break

            if id_col:
                # 设置该列所有单元格的格式为文本（包括表头和数据）
                for row_idx in range(1, ws.max_row + 1):  # 包括表头
                    cell = ws.cell(row=row_idx, column=id_col)
                    if cell.value is not None:
                        cell.value = str(cell.value)
                    cell.number_format = '@'  # '@' 表示文本格式

        if not silent:
            print(f"  保存 {len(all_records)} 条记录到 {output_file}")

        return True

    except Exception as e:
        if not silent:
            print(f"  [Error] 保存记录到Excel文件失败: {e}")
        return False

def flatten_live_data(live_object):
    """将直播列表数据展平为标准格式

    Args:
        live_object: 直播对象数据字典，包含直播的基本信息和统计数据

    Returns:
        dict: 展平后的直播数据，包含liveObjectId、直播信息、直播时长、观看人数等字段
    """
    live_stats = live_object.get('liveStats', {})

    # 只保留需要的字段，按照表头顺序
    flat_data = {
        'liveObjectId': str(live_object.get('liveObjectId')),  # 转换为字符串
        '直播信息': live_object.get('description'),
        '直播时长': live_object.get('liveStats', {}).get('liveDurationInSeconds', 0),
        '观看人数': live_object.get('liveStats', {}).get('totalAudienceCount', 0),
        '最高在线': live_object.get('maxOnlineCount', 0),
        '总热度': live_object.get('hotQuota', 0),
        '成交金额': live_object.get('payedGmv', '0'),
    }

    return flat_data

def backup_file(file_path):
    """备份文件，如果文件存在则复制并重命名为_backup

    Args:
        file_path: 要备份的文件路径
    """
    if os.path.exists(file_path):
        # 生成备份文件名
        if file_path.endswith('.xlsx'):
            backup_path = file_path[:-5] + '_backup.xlsx'
        else:
            backup_path = file_path + '_backup'

        try:
            shutil.copy2(file_path, backup_path)
            print(f"[备份] 已备份文件: {backup_path}")
        except Exception as e:
            print(f"[Warning] 备份文件失败: {e}")


def download_detail_data(output_file='xlsx2.xlsx', user_data_dir='./browser_data'):
    """下载预约数据（接口2）"""
    return download_api_data(
        output_file=output_file,
        data_type_name='预约数据',
        fetch_func=fetch_live_single_data,
        flatten_func=flatten_live_single_data,
        sheet_name='预约数据',
        id_column_name='liveObjectId',
        user_data_dir=user_data_dir
    )

def download_product_data(output_file='xlsx3.xlsx', user_data_dir='./browser_data'):
    """下载直播带货商品SPU数据（接口3）

    Args:
        output_file: 输出Excel文件路径，默认 'xlsx3.xlsx'
        user_data_dir: 浏览器数据目录，用于获取cookies和headers

    Returns:
        bool: 下载是否成功
    """
    # 备份旧文件
    backup_file(output_file)

    print(f"开始下载直播商品SPU数据...")
    print(f"输出文件: {output_file}")

    all_records = []

    # 读取 liveObjectId 列表（来自 xlsx1.xlsx）
    try:
        # 尝试读取新工作表名称，如果不存在则回退到旧名称（向后兼容）
        try:
            df_list = pd.read_excel('xlsx1.xlsx', sheet_name='列表数据')
        except ValueError:
            # 如果新工作表名称不存在，尝试旧的工作表名称
            df_list = pd.read_excel('xlsx1.xlsx', sheet_name='直播数据')
        live_ids = [str(live_id) for live_id in df_list['liveObjectId'].tolist()]
    except Exception as e:
        print(f"读取 xlsx1.xlsx 失败: {e}")
        return False

    # 尝试从浏览器会话获取 headers/cookies（只做一次）
    browser_headers, browser_cookies = get_browser_session_cookies_and_headers(
        user_data_dir=user_data_dir,
        url=URL_DETAIL  # 使用接口2的URL来获取cookies
    )
    if browser_headers or browser_cookies:
        print("已从浏览器会话获取 cookies/headers，将用于接口请求")
    else:
        browser_headers, browser_cookies = None, None

    for idx, live_id in enumerate(live_ids, 1):
        print(f"[{idx}/{len(live_ids)}] 获取 {live_id} 的带货商品的数据...")
        data = fetch_spu_data(live_id, headers=browser_headers, cookies=browser_cookies)
        if data is None:
            print(f"  警告: 未获取到数据，保存空记录")
            rec = {'liveObjectId': str(live_id)}
            all_records.append(rec)
        else:
            rec_list = flatten_spu_data(live_id, data)
            all_records.extend(rec_list)

        # 每次请求间隔 1 秒，避免过快
        time.sleep(1)

        # 每 50 条实时保存一次，防止意外中断丢失数据
        if idx % 50 == 0:
            save_records_to_excel_file(output_file, all_records, sheet_name='产品数据', id_column_name='liveObjectId', silent=True)

    # 最终保存
    success = save_records_to_excel_file(output_file, all_records, sheet_name='产品数据', id_column_name='liveObjectId', silent=False)
    if success:
        print(f"直播商品SPU数据已保存到 {output_file}，共 {len(all_records)} 条记录")
        return True
    else:
        print("保存直播商品SPU数据失败")
        return False


def fetch_live_single_data(live_object_id, headers=None, cookies=None, timeout=10):
    """调用接口2，获取指定 liveObjectId 的预约数据汇总，返回 data 字典或 None"""
    payload = {
        "liveObjectId": str(live_object_id),
        "timestamp": str(int(time.time() * 1000)),
        "_log_finder_uin": None,
        "_log_finder_id": "v2_060000231003b20faec8c5e58e18c6d4c605ed31b0777108d955d806e1454ae22f3ddeb0baf6@finder",
        "rawKeyBuff": None,
        "pluginSessionId": None,
        "scene": 7,
        "reqScene": 7
    }

    request_headers = headers if headers is not None and headers else HEADERS
    request_cookies = cookies if cookies is not None and cookies else COOKIES_DICT

    try:
        resp = requests.post(URL_DETAIL, json=payload, headers=request_headers, cookies=request_cookies, timeout=timeout)
        resp.raise_for_status()
        j = resp.json()
        if j.get('errCode') == 0:
            return j.get('data', {})
        else:
            print(f"接口2返回错误: {j.get('errMsg')}")
            return None
    except Exception as e:
        print(f"请求接口2失败: {e}")
        return None


def fetch_ec_summary(live_object_id, headers=None, cookies=None, timeout=10):
    """调用接口4，获取指定 liveObjectId 的带货数据的整体转换数据，返回 data 字典或 None"""
    payload = {
        "liveObjectId": str(live_object_id),
        "timestamp": str(int(time.time() * 1000)),
        "_log_finder_uin": None,
        "_log_finder_id": None,
        "rawKeyBuff": None,
        "pluginSessionId": None,
        "scene": 7,
        "reqScene": 7
    }

    request_headers = headers if headers is not None and headers else HEADERS
    request_cookies = cookies if cookies is not None and cookies else COOKIES_DICT

    try:
        resp = requests.post(URL_EC_SUMMARY, json=payload, headers=request_headers, cookies=request_cookies, timeout=timeout)
        resp.raise_for_status()
        j = resp.json()
        if j.get('errCode') == 0:
            return j.get('data', {})
        else:
            print(f"接口4返回错误: {j.get('errMsg')}")
            return None
    except Exception as e:
        print(f"请求接口4失败: {e}")
        return None


def fetch_spu_data(live_object_id, headers=None, cookies=None, timeout=10):
    """调用接口3，获取指定 liveObjectId 的带货商品的数据，返回 data 字典或 None"""
    payload = {
        "liveObjectId": str(live_object_id),
        "offset": 0,
        "limit": 15,
        "spuType": 0,
        "spuThreshold": {
            "lowStock": "10",
            "unpaidOrder": "10",
            "newBuyerConv": "10"
        },
        "spuSrc": 0,
        "fieldList": [
            "stock", "create_pv", "pay_pv", "gmv", "clk_pay_ratio", "create_uv",
            "pay_uv", "new_customer_pay_pv", "no_finish_pv", "share_uv", "exp_uv",
            "exp_pv", "clk_uv", "clk_pv", "exp_clk_ratio", "clk_pay_ratio_pv",
            "new_customer_conversion_rate", "id", "explanation_count",
            "new_customer_conversion_rate_pv", "refund_rate", "refund_uv",
            "refund_pv", "refund_amount"
        ],
        "timestamp": str(int(time.time() * 1000)),
        "_log_finder_uin": None,
        "_log_finder_id": "v2_060000231003b20faec8c5e58e18c6d4c605ed31b0777108d955d806e1454ae22f3ddeb0baf6@finder",
        "rawKeyBuff": None,
        "pluginSessionId": None,
        "scene": 7,
        "reqScene": 7
    }

    request_headers = headers if headers is not None and headers else HEADERS
    request_cookies = cookies if cookies is not None and cookies else COOKIES_DICT

    try:
        resp = requests.post(URL_PRODUCT, json=payload, headers=request_headers, cookies=request_cookies, timeout=timeout)
        resp.raise_for_status()
        j = resp.json()
        if j.get('errCode') == 0:
            return j.get('data', {})
        else:
            print(f"接口3返回错误: {j.get('errMsg')}")
            return None
    except Exception as e:
        print(f"请求接口3失败: {e}")
        return None


def fetch_live_diagnostic_data(live_object_id, headers=None, cookies=None, timeout=10):
    """调用接口5，获取指定 liveObjectId 的数据增强诊断数据，返回 data 字典或 None"""
    payload = {
        "objectId": str(live_object_id),
        "timestamp": str(int(time.time() * 1000)),
        "_log_finder_uin": None,
        "_log_finder_id": "v2_060000231003b20faec8c5e58e18c6d4c605ed31b0777108d955d806e1454ae22f3ddeb0baf6@finder",
        "rawKeyBuff": None,
        "pluginSessionId": None,
        "scene": 7,
        "reqScene": 7
    }

    request_headers = headers if headers is not None and headers else HEADERS
    request_cookies = cookies if cookies is not None and cookies else COOKIES_DICT

    print(f"  [接口5] 请求参数: objectId={live_object_id}")

    try:
        resp = requests.post(URL_DIAGNOSTIC, json=payload, headers=request_headers, cookies=request_cookies, timeout=timeout)
        resp.raise_for_status()
        j = resp.json()
        print(f"  [接口5] 响应状态码: {resp.status_code}, errCode: {j.get('errCode')}")
        if j.get('errCode') == 0:
            data = j.get('data', {})
            if data and 'newWatchPvPromotion' in data:
                promotion_value = data['newWatchPvPromotion'].get('value', 'N/A')
                print(f"  [接口5] 获取到newWatchPvPromotion: {promotion_value}")
            return data
        else:
            print(f"接口5返回错误: {j.get('errMsg')}")
            return None
    except Exception as e:
        print(f"请求接口5失败: {e}")
        return None


def flatten_live_single_data(live_object_id, single_data):
    """将接口2的 data 展平为一条记录（dict），并确保所有值为字符串"""
    if single_data is None:
        return None

    flat = {'liveObjectId': str(live_object_id)}

    # 处理预约通知用户数和比率
    flat['reserveNoticeUserCount'] = str(single_data.get('reserveNoticeUserCount', ''))
    flat['reserveNoticeJoinliveRatio'] = str(single_data.get('reserveNoticeJoinliveRatio', ''))

    # 处理场景数组，将每个场景的数据作为单独字段
    # 遍历data中的所有项，找出数组类型的项（即场景数据）
    for key, value in single_data.items():
        if isinstance(value, list):
            # 这是场景数据数组
            for item in value:
                if isinstance(item, dict) and 'scene' in item and 'reserveNoticeUserCount' in item:
                    scene = item['scene']
                    count = item['reserveNoticeUserCount']
                    flat[f'scene_{scene}_reserveNoticeUserCount'] = str(count)

    return flat


def flatten_ec_summary(live_object_id, ec_data):
    """将接口4的 data 展平为一条记录（dict），并确保所有值为字符串"""
    if ec_data is None:
        return None

    flat = {'liveObjectId': str(live_object_id)}
    # 将 data 中的键值全部转为字符串并加入
    for k, v in ec_data.items():
        # 保持原始键名，值转换为字符串（None -> ''）
        if v is None:
            flat[k] = ''
        elif isinstance(v, (int, float, bool)):
            flat[k] = str(v)
        else:
            flat[k] = str(v).strip()

    return flat


def flatten_spu_data(live_object_id, spu_data):
    """将接口3的 data 展平为多条记录（list），每条记录的第一列是liveObjectId"""
    if spu_data is None:
        return []

    spu_data_list = spu_data.get('spuDataList', [])

    if not spu_data_list:
        # 如果没有数据，至少返回一条包含liveObjectId的记录
        return [{'liveObjectId': str(live_object_id)}]

    flattened_data = []

    for spu_item in spu_data_list:
        flat_record = {'liveObjectId': str(live_object_id)}

        # 处理baseData字段
        base_data = spu_item.get('baseData', {})
        if base_data:
            flat_record['srcSpuId'] = str(base_data.get('srcSpuId', ''))
            flat_record['spuId'] = str(base_data.get('spuId', ''))
            flat_record['src'] = str(base_data.get('src', ''))
            flat_record['spuName'] = str(base_data.get('spuName', ''))
            flat_record['thumbUrl'] = str(base_data.get('thumbUrl', ''))
            flat_record['price'] = str(base_data.get('price', ''))
            flat_record['srcName'] = str(base_data.get('srcName', ''))
            flat_record['baseStock'] = str(base_data.get('stock', ''))  # 重命名避免冲突

        # 处理其他字段
        for key, value in spu_item.items():
            if key not in ['baseData']:  # baseData已单独处理
                if value is None:
                    flat_record[key] = ''
                elif isinstance(value, (int, float, bool)):
                    flat_record[key] = str(value)
                else:
                    flat_record[key] = str(value).strip()

        flattened_data.append(flat_record)

    return flattened_data


def flatten_live_diagnostic_data(live_object_id, diagnostic_data):
    """将接口5的 data 展平为一条记录（dict），主要提取newWatchPvPromotion字段"""
    if diagnostic_data is None:
        return None

    flat = {'liveObjectId': str(live_object_id)}

    # 主要提取newWatchPvPromotion字段
    new_watch_pv_promotion = diagnostic_data.get('newWatchPvPromotion', {})
    if new_watch_pv_promotion:
        flat['newWatchPvPromotion'] = str(new_watch_pv_promotion.get('value', ''))
    else:
        flat['newWatchPvPromotion'] = ''

    return flat


def download_ec_summary(output_file='xlsx4.xlsx', user_data_dir='./browser_data'):
    """下载带货数据的整体转换数据（接口4）"""
    return download_api_data(
        output_file=output_file,
        data_type_name='带货数据的整体转换数据',
        fetch_func=fetch_ec_summary,
        flatten_func=flatten_ec_summary,
        sheet_name='EC汇总',
        id_column_name='liveObjectId',
        user_data_dir=user_data_dir
    )


def download_live_diagnostic_data(input_file='xlsx1.xlsx', user_data_dir='./browser_data'):
    """下载数据增强诊断数据（接口5），并将数据插入到xlsx1.xlsx的newWatchPvPromotion列中

    Args:
        input_file: 输入的xlsx1.xlsx文件路径
        user_data_dir: 浏览器数据目录，用于获取cookies和headers

    Returns:
        bool: 下载是否成功
    """
    # 备份原文件
    backup_file(input_file)

    print(f"开始下载数据增强诊断数据...")
    print(f"输入文件: {input_file}")

    try:
        # 读取现有的xlsx1.xlsx文件
        try:
            df = pd.read_excel(input_file, sheet_name='列表数据')
        except ValueError:
            # 如果新工作表名称不存在，尝试旧的工作表名称
            df = pd.read_excel(input_file, sheet_name='直播数据')

        # 确保liveObjectId列存在
        if 'liveObjectId' not in df.columns:
            print("错误: xlsx1.xlsx中没有找到liveObjectId列")
            return False

        live_ids = [str(live_id) for live_id in df['liveObjectId'].tolist()]
        print(f"找到 {len(live_ids)} 个直播ID需要处理")

        # 尝试从浏览器会话获取 headers/cookies（只做一次）
        browser_headers, browser_cookies = get_browser_session_cookies_and_headers(
            user_data_dir=user_data_dir,
            url=URL_DIAGNOSTIC
        )
        if browser_headers or browser_cookies:
            print("已从浏览器会话获取 cookies/headers，将用于接口请求")
        else:
            browser_headers, browser_cookies = None, None

        # 为每个liveObjectId获取诊断数据
        new_watch_pv_promotion_values = []

        for idx, live_id in enumerate(live_ids, 1):
            print(f"[{idx}/{len(live_ids)}] 获取 {live_id} 的数据增强诊断数据...")
            data = fetch_live_diagnostic_data(live_id, headers=browser_headers, cookies=browser_cookies)

            if data is None:
                print(f"  警告: 未获取到 {live_id} 的数据，使用空值")
                new_watch_pv_promotion_values.append('')
            else:
                flattened = flatten_live_diagnostic_data(live_id, data)
                if flattened and 'newWatchPvPromotion' in flattened:
                    value = flattened['newWatchPvPromotion']
                    print(f"  获取到newWatchPvPromotion: {value}")
                    new_watch_pv_promotion_values.append(value)
                else:
                    print(f"  警告: {live_id} 的数据格式异常，使用空值")
                    new_watch_pv_promotion_values.append('')

            # 每次请求间隔 1 秒，避免过快
            time.sleep(1)

        # 将新数据添加到DataFrame
        df['newWatchPvPromotion'] = new_watch_pv_promotion_values

        # 保存回Excel文件
        with pd.ExcelWriter(input_file, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False, sheet_name='列表数据')

            # 获取工作表并设置liveObjectId列为文本格式
            ws = writer.sheets['列表数据']
            # 查找liveObjectId列的索引
            header_row = 1
            id_col = None
            for col_idx, cell in enumerate(ws[header_row], 1):
                if cell.value == 'liveObjectId':
                    id_col = col_idx
                    break

            if id_col:
                # 设置该列所有单元格的格式为文本
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=id_col)
                    if cell.value is not None:
                        cell.value = str(cell.value)
                    cell.number_format = '@'  # '@' 表示文本格式

        print(f"数据增强诊断数据已更新到 {input_file}，共处理 {len(live_ids)} 条记录")
        return True

    except Exception as e:
        print(f"[Error] 下载数据增强诊断数据失败: {e}")
        return False

def check_login_status(user_data_dir='./browser_data'):
    """检查登录状态，询问用户是否已登录，如果未登录则使用 Playwright 打开登录页面"""
    login_url = 'https://channels.weixin.qq.com/login.html'
    
    print("=" * 60)
    print("登录状态检查")
    print("=" * 60)
    
    while True:
        user_input = input("您是否已登录微信视频号助手？(y/n): ").strip().lower()
        
        if user_input in ['y', 'yes', '是', 'Y']:
            print("已确认登录状态，开始执行程序...")
            return True
        elif user_input in ['n', 'no', '否', 'N']:
            print(f"\n请先登录后再运行程序。")
            print(f"登录地址: {login_url}")
            print("\n正在使用 Playwright 打开登录页面...")
            
            # 使用 Playwright 打开登录页面
            playwright = sync_playwright().start()
            try:
                # 使用持久化上下文，这样登录状态会被保存
                context = playwright.chromium.launch_persistent_context(
                    user_data_dir=user_data_dir,
                    headless=False,
                    viewport={'width': 1920, 'height': 1080},
                    locale='zh-CN',
                    timezone_id='Asia/Shanghai',
                    args=[
                        '--disable-blink-features=AutomationControlled',
                        '--window-size=1920,1080',
                    ]
                )
                
                # 获取第一个页面或创建新页面
                pages = context.pages
                if pages:
                    page = pages[0]
                else:
                    page = context.new_page()
                
                # 打开登录页面
                page.goto(login_url)
                print("登录页面已打开，请在浏览器中完成登录。")
                print("登录完成后，请关闭浏览器窗口，然后重新运行此程序。")
                print("\n按 Enter 键继续...")
                input()
                
                # 关闭浏览器上下文
                context.close()
                playwright.stop()
            except Exception as e:
                print(f"打开登录页面失败: {e}")
                print(f"请手动访问: {login_url}")
                try:
                    playwright.stop()
                except:
                    pass
            
            print("\n程序已退出。请登录后重新运行。")
            return False
        else:
            print("请输入 y 或 n（是/否）")

def download_api_data(
    output_file,
    data_type_name,
    fetch_func,
    flatten_func,
    sheet_name,
    id_column_name,
    user_data_dir='./browser_data',
    is_batch_request=False,
    batch_params=None
):
    """
    统一的API数据下载函数

    Args:
        output_file: 输出文件名
        data_type_name: 数据类型名称（用于日志输出）
        fetch_func: 数据获取函数
        flatten_func: 数据展平函数
        sheet_name: Excel工作表名称
        id_column_name: ID列名称
        user_data_dir: 浏览器数据目录
        is_batch_request: 是否为批量请求（如接口1的分页）
        batch_params: 批量请求的参数（仅当is_batch_request=True时使用）
    """
    # 备份旧文件
    backup_file(output_file)

    print(f"开始下载{data_type_name}...")
    print(f"输出文件: {output_file}")

    all_records = []

    # 对于批量请求（接口1），不需要读取xlsx1.xlsx，直接进行批量获取
    if is_batch_request:
        # 尝试从浏览器会话获取 headers/cookies（只做一次）
        browser_headers, browser_cookies = get_browser_session_cookies_and_headers(
            user_data_dir=user_data_dir,
            url=URL_LIST
        )
        if browser_headers or browser_cookies:
            print("已从浏览器会话获取 cookies/headers，将用于接口请求")
        else:
            browser_headers, browser_cookies = None, None

        # 批量请求处理（接口1）
        start_date = batch_params.get('start_date') if batch_params else None
        end_date = batch_params.get('end_date') if batch_params else None
        start_time, end_time = get_time_range_for_half_year(start_date, end_date)

        start_datetime = datetime.fromtimestamp(start_time)
        end_datetime = datetime.fromtimestamp(end_time)
        print(f"时间范围: {start_datetime.strftime('%Y-%m-%d %H:%M:%S')} 到 {end_datetime.strftime('%Y-%m-%d %H:%M:%S')}")
        if start_date or end_date:
            print(f"自定义时间范围: {start_date or '默认'} 到 {end_date or '当前时间'}")

        current_page = 1
        page_size = batch_params.get('page_size', 50) if batch_params else 50
        total_count = None

        while True:
            print(f"正在下载第 {current_page} 页...")

            result = fetch_func(
                page_size=page_size,
                current_page=current_page,
                start_time=start_time,
                end_time=end_time,
                headers=browser_headers,
                cookies=browser_cookies
            )

            if result is None:
                print(f"第 {current_page} 页下载失败，停止")
                break

            data_list = result.get('liveObjectList', [])

            if not data_list:
                print(f"第 {current_page} 页无数据，下载完成")
                break

            # 展平数据并添加到列表
            for data_obj in data_list:
                flat_obj = flatten_func(data_obj)
                all_records.append(flat_obj)

            # 获取总数
            if total_count is None:
                total_count = result.get('totalLiveCount', 0)
                print(f"总共有 {total_count} 条数据")

            print(f"已下载 {len(all_records)} 条数据")

            # 检查是否已下载所有数据
            if len(all_records) >= total_count:
                print(f"已获取所有 {total_count} 条数据")
                break

            current_page += 1
            time.sleep(1)  # 暂停1秒，避免请求过于频繁
    else:
        # 单条请求处理（接口2、4）- 需要先读取xlsx1.xlsx获取liveObjectId列表
        # 读取 liveObjectId 列表（来自 xlsx1.xlsx）
        try:
            # 尝试读取新工作表名称，如果不存在则回退到旧名称（向后兼容）
            try:
                df_list = pd.read_excel('xlsx1.xlsx', sheet_name='列表数据')
            except ValueError:
                # 如果新工作表名称不存在，尝试旧的工作表名称
                df_list = pd.read_excel('xlsx1.xlsx', sheet_name='直播数据')
            live_ids = [str(live_id) for live_id in df_list['liveObjectId'].tolist()]
        except Exception as e:
            print(f"读取 xlsx1.xlsx 失败: {e}")
            return False

        # 尝试从浏览器会话获取 headers/cookies（只做一次）
        browser_headers, browser_cookies = get_browser_session_cookies_and_headers(
            user_data_dir=user_data_dir,
            url=URL_LIST if data_type_name == '列表数据' else None
        )
        if browser_headers or browser_cookies:
            print("已从浏览器会话获取 cookies/headers，将用于接口请求")
        else:
            browser_headers, browser_cookies = None, None

        for idx, live_id in enumerate(live_ids, 1):
            print(f"[{idx}/{len(live_ids)}] 获取 {live_id} 的{data_type_name}...")
            data = fetch_func(live_id, headers=browser_headers, cookies=browser_cookies)
            if data is None:
                print(f"  警告: 未获取到数据，保存空记录")
                rec = {id_column_name: str(live_id)}
                all_records.append(rec)
            else:
                rec = flatten_func(live_id, data)
                all_records.append(rec)

            # 每次请求间隔 1 秒，避免过快
            time.sleep(1)

            # 每 50 条实时保存一次，防止意外中断丢失数据
            if idx % 50 == 0:
                save_records_to_excel_file(output_file, all_records, sheet_name=sheet_name, id_column_name=id_column_name, silent=True)

    # 最终保存
    success = save_records_to_excel_file(output_file, all_records, sheet_name=sheet_name, id_column_name=id_column_name, silent=False)
    if success:
        print(f"{data_type_name}已保存到 {output_file}，共 {len(all_records)} 条记录")
        return True
    else:
        print(f"保存{data_type_name}失败")
        return False

def download_half_year_data(output_file='xlsx1.xlsx', user_data_dir='./browser_data', start_date=None, end_date=None):
    """下载列表数据（接口1）

    Args:
        output_file: 输出文件名
        user_data_dir: 浏览器数据目录
        start_date: 开始日期，格式为 'YYYY-MM-DD'，默认为今年1月1号
        end_date: 结束日期，格式为 'YYYY-MM-DD'，默认为当前日期

    Examples:
        # 使用默认时间范围（今年1月1号到当前时间）
        download_half_year_data()

        # 指定时间范围
        download_half_year_data(start_date='2024-01-01', end_date='2024-12-31')

        # 只指定开始日期，结束日期使用当前时间
        download_half_year_data(start_date='2024-06-01')

        # 只指定结束日期，开始日期使用默认值（今年1月1号）
        download_half_year_data(end_date='2024-12-31')
    """
    return download_api_data(
        output_file=output_file,
        data_type_name='列表数据',
        fetch_func=fetch_live_data,
        flatten_func=flatten_live_data,
        sheet_name='列表数据',
        id_column_name='liveObjectId',
        user_data_dir=user_data_dir,
        is_batch_request=True,
        batch_params={
            'page_size': 50,
            'start_date': start_date,
            'end_date': end_date
        }
    )

if __name__ == '__main__':

    # 检查登录状态
    if not check_login_status():
        exit(0)

    print("\n" + "=" * 60)
    print("开始执行数据下载任务")
    print("=" * 60 + "\n")

    # 下载列表数据（接口1）
    print("正在下载列表数据（接口1）...")
    success1 = download_half_year_data(start_date='2025-01-01')  # 使用默认时间范围（今年1月1号到当前时间）
    # download_half_year_data(start_date='2024-01-01', end_date='2024-12-31')  # 指定时间范围

    if success1:
        print("\n" + "=" * 60)
        print("列表数据下载完成，开始下载其他接口数据")
        print("=" * 60 + "\n")

        # 下载预约数据（接口2）- 下载全部数据
        # 输出文件名为 xlsx2.xlsx
        print("正在下载预约数据（接口2）...")
        download_detail_data(output_file='预约数据.xlsx')

        # 下载带货商品的数据（接口3）- 下载全部数据
        # 输出文件名为 xlsx3.xlsx
        print("正在下载带货商品的数据（接口3）...")
        download_product_data(output_file='带货商品数据.xlsx')

        # 下载带货数据的整体转换数据（接口4）
        # 输出文件名为 xlsx4_YYYYMMDDHHMMSS.xlsx
        print("正在下载带货数据的整体转换数据（接口4）...")
        download_ec_summary(output_file='整体转换.xlsx')

        # 下载数据增强诊断数据（接口5）- 更新xlsx1.xlsx文件
        print("正在下载数据增强诊断数据（接口5）...")
        download_live_diagnostic_data(input_file='xlsx1.xlsx')

        print("\n" + "=" * 60)
        print("所有接口数据下载完成！")
        print("=" * 60 + "\n")
    else:
        print("\n" + "=" * 60)
        print("列表数据下载失败，跳过其他接口的下载")
        print("=" * 60 + "\n")
